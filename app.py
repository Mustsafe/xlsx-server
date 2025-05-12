from flask import Flask, request, jsonify, Response
import pandas as pd
import os
import re
import json
import difflib
import requests
from bs4 import BeautifulSoup
from io import BytesIO
from typing import List
from urllib.parse import quote
from datetime import datetime, timedelta
from dateutil import parser
import openai
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 로거 설정
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
openai.api_key = os.getenv("OPENAI_API_KEY")
DATA_DIR = "./data"
os.makedirs(DATA_DIR, exist_ok=True)

# 유틸 함수

def sanitize(text: str) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", text.lower())

# 템플릿명 alias 매핑

def build_alias_map(templates: List[str]) -> dict:
    alias = {}
    for tpl in templates:
        key = sanitize(tpl)
        alias[key] = tpl
    return alias

# 키워드로 템플릿 resolve

def resolve_keyword(raw: str, templates: List[str], alias_map: dict, freq: dict) -> str:
    query = re.sub(r"\s*(?:양식|서식)(?:을|를)?$", "", raw.strip(), flags=re.IGNORECASE)
    key = sanitize(query)
    if key in alias_map:
        return alias_map[key]
    norms = [sanitize(t) for t in templates]
    match = difflib.get_close_matches(key, norms, n=1, cutoff=0.6)
    if match:
        return templates[norms.index(match[0])]
    raise ValueError(f"템플릿 '{raw}'을(를) 찾을 수 없습니다.")

# JSON 배열 행 분해 (BOM·스마트 따옴표 처리 포함)

def explode_json_rows(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        raw_text = row.get("작성 양식", "")
        text = raw_text.lstrip("\ufeff").strip()
        text = text.replace("“", '"').replace("”", '"')
        try:
            arr = json.loads(text) if text.startswith("[") else None
        except json.JSONDecodeError:
            arr = None
        if isinstance(arr, list):
            for elem in arr:
                new = row.copy()
                new["작업 항목"] = elem.get("항목", "")
                new["작성 양식"] = elem.get("내용", elem.get("세부사항", ""))
                records.append(new)
        else:
            records.append(row)
    return pd.DataFrame(records)

@app.route("/create_xlsx", methods=["GET"])
def create_xlsx():
    try:
        raw = request.args.get("template", "")
        path = os.path.join(DATA_DIR, "통합_노지파일.csv")
        if not os.path.exists(path):
            return jsonify(error="통합 CSV 파일이 없습니다."), 404

        df = pd.read_csv(path, encoding="utf-8-sig")
        templates = df["템플릿명"].dropna().unique().tolist()
        alias_map = build_alias_map(templates)\n        freq = df["템플릿명"].value_counts().to_dict()

        try:
            tpl = resolve_keyword(raw, templates, alias_map, freq)
            out_df = df[df["템플릿명"] == tpl][["작업 항목","작성 양식","실무 예시 1","실무 예시 2"]].copy()
        except ValueError:
            tpl = raw
            system = {"role":"system","content":"산업안전 문서 템플릿 전문가입니다. JSON 배열을 5개 이상 생성하세요."}
            user = {"role":"user","content":f"템플릿명 '{raw}'의 기본 양식을 JSON 배열로 주세요."}
            resp = openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[system, user],
                max_tokens=800,
                temperature=0.7
            )
            try:
                data = json.loads(resp.choices[0].message.content)
                out_df = pd.DataFrame(data)
            except:
                out_df = pd.DataFrame([{
                    "작업 항목": raw,
                    "작성 양식": resp.choices[0].message.content.replace("\n"," "),
                    "실무 예시 1": "",
                    "실무 예시 2": ""
                }])

        # JSON 배열 분해
        out_df = explode_json_rows(out_df)

        # AI 동적 고도화
        for idx, row in out_df.iterrows():
            base = row["작성 양식"]
            if isinstance(base, str) and len(base.splitlines()) < 3:
                sys_msg = {"role":"system","content":"5~8개 점검 리스트를 JSON 배열로 생성하세요."}
                usr_msg = {"role":"user","content":json.dumps({"base": base})}
                try:
                    r = openai.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[sys_msg, usr_msg],
                        max_tokens=300,
                        temperature=0.7
                    )
                    items = json.loads(r.choices[0].message.content)
                    if isinstance(items, list):
                        out_df.at[idx, "작성 양식"] = "\n".join(items)
                except:
                    pass
            for ex in ["실무 예시 1","실무 예시 2"]:
                ex_base = row.get(ex, "")
                if ex_base:
                    sysg = {"role":"system","content":"구체적 현장 사례 한 문장으로 설명하세요."}
                    usrg = {"role":"user","content":json.dumps({"base": ex_base})}
                    try:
                        rr = openai.chat.completions.create(
                            model="gpt-4o-mini",
                            messages=[sysg, usrg],
                            max_tokens=100,
                            temperature=0.7
                        )
                        out_df.at[idx, ex] = rr.choices[0].message.content.strip()
                    except:
                        pass

        # 순서 재정렬
        order = ["📋 작업 절차","💡 실무 가이드","✅ 체크리스트","🛠️ 유지보수 포인트","📎 출처"]
        out_df["_order"] = out_df["작업 항목"].apply(lambda x: order.index(x) if x in order else 99)
        out_df = out_df.sort_values("_order").drop(columns=["_order"])

        # 엑셀 생성 & 포맷
        wb = Workbook()
        ws = wb.active
        headers = ["작업 항목","작성 양식","실무 예시 1","실무 예시 2"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in out_df.itertuples(index=False):
            ws.append(row)
        for i, col in enumerate(ws.columns, 1):
            mx = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(i)].width = min(mx+2,60)
            for c in col[1:]:
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{tpl}.xlsx" if tpl else f"{raw}.xlsx"
        disp = quote(filename)
        return Response(
            buf.read(),
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": f"attachment; filename*=UTF-8''{disp}",
                "Cache-Control": "public, max-age=3600"
            }
        )
    except Exception as e:
        logger.exception("create_xlsx error")
        return jsonify(error=f"서버 오류: {e}"), 500

# 뉴스 크롤링 & 렌더링 로직

def fetch_safetynews_article_content(url):
    try:
        r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        node = soup.select_one("div#article-view-content-div")
        return node.get_text("\n").strip() if node else ""
    except:
        return ""

def crawl_naver_news():
    base = "https://openapi.naver.com/v1/search/news.json"
    headers = {"X-Naver-Client-Id":NAVER_CLIENT_ID, "X-Naver-Client-Secret":NAVER_CLIENT_SECRET}
    kws = ["건설 사고","추락 사고","끼임 사고","질식 사고","폭발 사고","산업재해","산업안전"]
    out=[]
    for kw in kws:
        r = requests.get(base, headers=headers, params={"query":kw, "display":2, "sort":"date"}, timeout=10)
        if r.status_code==200:
            for item in r.json().get("items",[]):
                title = BeautifulSoup(item["title"],"html.parser").get_text()
                desc  = BeautifulSoup(item["description"],"html.parser").get_text()
